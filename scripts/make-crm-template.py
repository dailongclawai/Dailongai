# Sinh file Excel mẫu nhập khách hàng CRM — MỘT chuẩn dùng chung web + bot Telegram.
# Đẹp như bản bot cũ nhưng an toàn: chú thích cột nằm ở tooltip tiêu đề (không có
# dòng chữ lẫn vào sheet dữ liệu), Nguồn + Trạng thái có dropdown, ví dụ nằm ở
# sheet Hướng dẫn.
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ORANGE = 'FF5625'
DARK = '1F2937'
GRAY = 'F3F4F6'

HEADERS = [
    ('Tên khách hàng *', 24, 'BẮT BUỘC. Thiếu tên là dòng đó bị bỏ qua khi Nhập.'),
    ('Điện thoại', 14, 'Nên có. CRM dùng số này để phát hiện khách trùng.'),
    ('Zalo', 14, 'Để trống nếu trùng số điện thoại.'),
    ('Tỉnh / Thành phố', 18, 'Ví dụ: Hà Nội'),
    ('Địa chỉ', 30, 'Số nhà, đường, quận/huyện.'),
    ('Nguồn', 14, 'Chọn trong danh sách xổ xuống. Nghĩa từng mã xem sheet Hướng dẫn.'),
    ('Trạng thái', 20, 'Chọn trong danh sách xổ xuống. Để trống thì vào hệ thống là "Mới tiếp nhận".'),
    ('Số lượng máy', 13, 'Số máy khách định mua (từ 1 trở lên). Có số máy sẽ tự lập cơ hội. Chưa rõ thì để trống.'),
    ('Ghi chú', 34, 'Nhu cầu, tình trạng, lịch hẹn…'),
    ('Email', 24, 'Không bắt buộc.'),
]

SOURCES = [
    ('website', 'Website dailongai.com'),
    ('zalo', 'Zalo'),
    ('facebook', 'Facebook'),
    ('google_ads', 'Quảng cáo Google'),
    ('tiktok', 'TikTok'),
    ('referral', 'Khách/đối tác giới thiệu'),
    ('hotline', 'Gọi vào hotline'),
    ('event', 'Sự kiện, hội thảo'),
    ('other', 'Nguồn khác'),
]

STAGES = ['Mới tiếp nhận', 'Đã liên hệ', 'Đang tư vấn', 'Đã trải nghiệm máy',
          'Đàm phán giá', 'Chốt đơn', 'Hoàn thành đơn', 'Không mua']

thin = Side(style='thin', color='D1D5DB')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ── Sheet 1: Khách hàng (sheet đầu — chức năng Nhập chỉ đọc sheet này) ─────────
ws = wb.active
ws.title = 'Khách hàng'
ws.freeze_panes = 'A2'
ws.row_dimensions[1].height = 24
for i, (title, width, tip) in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=i, value=title)
    c.font = Font(bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=ORANGE)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border
    c.comment = Comment(tip, 'CRM Đại Long', height=90, width=260)
    ws.column_dimensions[get_column_letter(i)].width = width

# Kẻ khung sẵn 30 dòng cho dễ điền
for r in range(2, 32):
    for i in range(1, len(HEADERS) + 1):
        ws.cell(row=r, column=i).border = border

dv_src = DataValidation(type='list', formula1='"' + ','.join(k for k, _ in SOURCES) + '"',
                        allow_blank=True, showErrorMessage=True,
                        errorTitle='Nguồn không hợp lệ', error='Chọn một mã trong danh sách xổ xuống.')
dv_stage = DataValidation(type='list', formula1='"' + ','.join(STAGES) + '"',
                          allow_blank=True, showErrorMessage=True,
                          errorTitle='Trạng thái không hợp lệ', error='Chọn đúng tên giai đoạn trong danh sách.')
dv_qty = DataValidation(type='whole', operator='greaterThanOrEqual', formula1='1',
                        allow_blank=True, showErrorMessage=True,
                        errorTitle='Số máy không hợp lệ', error='Nhập số nguyên từ 1 trở lên.')
ws.add_data_validation(dv_src); dv_src.add('F2:F501')
ws.add_data_validation(dv_stage); dv_stage.add('G2:G501')
ws.add_data_validation(dv_qty); dv_qty.add('H2:H501')

# ── Sheet 2: Hướng dẫn ────────────────────────────────────────────────────────
hd = wb.create_sheet('Hướng dẫn')
hd.column_dimensions['A'].width = 20
hd.column_dimensions['B'].width = 90
hd.sheet_view.showGridLines = False

def title_cell(row, text, size=12, color=ORANGE):
    c = hd.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=size, color=color)
    return c

def kv(row, a, b, head=False):
    ca = hd.cell(row=row, column=1, value=a)
    cb = hd.cell(row=row, column=2, value=b)
    for c in (ca, cb):
        c.border = border
        c.alignment = Alignment(vertical='top', wrap_text=True)
    if head:
        for c in (ca, cb):
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor=DARK)
    return row + 1

r = 1
title_cell(r, 'MẪU NHẬP KHÁCH HÀNG VÀO CRM dailongai.com', size=14); r += 2

title_cell(r, 'CÁCH DÙNG'); r += 1
for step in [
    '1. Điền khách vào sheet "Khách hàng" — mỗi dòng một khách, ĐỪNG chèn dòng chú thích.',
    '2. Rê chuột vào tiêu đề cột để xem giải thích; cột Nguồn và Trạng thái bấm vào ô là có danh sách xổ xuống.',
    '3. Lưu file rồi vào dailongai.com/portal/crm/accounts → nút "Nhập" → chọn file này.',
    '4. Hệ thống tự khớp cột, báo trước dòng trùng số điện thoại rồi mới nhập.',
]:
    hd.cell(row=r, column=1, value=step).alignment = Alignment(wrap_text=True)
    hd.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1
r += 1

title_cell(r, 'Ý NGHĨA TỪNG CỘT'); r += 1
r = kv(r, 'Cột', 'Giải thích', head=True)
for name, _, tip in HEADERS:
    r = kv(r, name, tip)
r += 1

title_cell(r, 'MÃ NGUỒN HỢP LỆ'); r += 1
r = kv(r, 'Mã', 'Nghĩa', head=True)
for k, v in SOURCES:
    r = kv(r, k, v)
r += 1

title_cell(r, 'TRẠNG THÁI HỢP LỆ'); r += 1
hd.cell(row=r, column=1, value=', '.join(STAGES)).alignment = Alignment(wrap_text=True)
hd.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
r += 1
hd.cell(row=r, column=1, value='Để trống = vào hệ thống ở bước "Mới tiếp nhận". Chọn "Không mua" thì hệ thống KHÔNG lập cơ hội cho dòng đó.')
hd.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
r += 2

title_cell(r, 'VÍ DỤ 2 DÒNG ĐIỀN ĐÚNG (chỉ để xem — đừng chép nguyên vào sheet Khách hàng)'); r += 1
ex_head = [h for h, _, _ in HEADERS]
ex_rows = [
    ['Nguyễn Văn A', '0912345678', '', 'Hà Nội', '12 Phố Huế, Hai Bà Trưng', 'zalo', 'Mới tiếp nhận', 1, 'Khách lẻ quan tâm máy laser', 'vana@gmail.com'],
    ['Phòng khám Đông y B', '0987654321', '', 'Thành phố Hồ Chí Minh', '45 Nguyễn Trãi, Quận 5', 'referral', 'Đang tư vấn', 2, 'Khách tổ chức — phòng khám', 'phongkhamb@gmail.com'],
]
for row_vals in [ex_head] + ex_rows:
    for i, v in enumerate(row_vals, start=1):
        c = hd.cell(row=r, column=i, value=v)
        c.border = border
        if row_vals is ex_head:
            c.font = Font(bold=True, color='FFFFFF', size=9)
            c.fill = PatternFill('solid', fgColor=DARK)
        else:
            c.font = Font(size=9)
    r += 1

out = 'public/templates/mau-nhap-khach-hang-crm.xlsx'
import os
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print('SAVED', out)
